from datetime import date, datetime, time

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import load_workbook

from .forms import (ComunicacionForm, ImportarComunicacionesForm,)
from .models import Comunicacion

def limpiar_texto(valor):
    if valor is None:
        return ""

    return str(valor).strip()


def convertir_fecha(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                str(valor).strip(),
                formato,
            ).date()

        except ValueError:
            continue

    return None


def convertir_hora(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.time()

    if isinstance(valor, time):
        return valor

    formatos = [
        "%H:%M:%S",
        "%H:%M",
        "%I:%M %p",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                str(valor).strip(),
                formato,
            ).time()

        except ValueError:
            continue

    return None

@login_required
def lista_comunicaciones(request):

    tipo = request.GET.get("tipo")

    comunicaciones = Comunicacion.objects.all().order_by(
        "-fecha_recibido",
        "-fecha_registro",
    )

    if tipo in ["RECIBIDA", "ENVIADA"]:
        comunicaciones = comunicaciones.filter(
            tipo=tipo
        )

    return render(
        request,
        "comunicaciones/lista.html",
        {
            "comunicaciones": comunicaciones,
            "tipo_actual": tipo,
        },
    )


@login_required
def crear_comunicacion(request):

    if request.method == "POST":

        form = ComunicacionForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            comunicacion = form.save()

            # Si esta comunicación responde a otra,
            # se marca la original como respondida.
            if comunicacion.respuesta_a:

                comunicacion_original = (
                    comunicacion.respuesta_a
                )

                comunicacion_original.estado = (
                    Comunicacion.EstadoComunicacion.RESPONDIDA
                )

                comunicacion_original.save()

            return redirect(
                "comunicaciones:lista"
            )

    else:

        form = ComunicacionForm()

    return render(
        request,
        "comunicaciones/formulario.html",
        {
            "form": form,
        },
    )


@login_required
def detalle_comunicacion(request, pk):

    comunicacion = get_object_or_404(
        Comunicacion,
        pk=pk,
    )

    respuestas = comunicacion.respuestas.all()

    return render(
        request,
        "comunicaciones/detalle.html",
        {
            "comunicacion": comunicacion,
            "respuestas": respuestas,
        },
    )


@login_required
def importar_excel(request):

    resultado = None

    if request.method == "POST":

        form = ImportarComunicacionesForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            archivo = form.cleaned_data["archivo_excel"]

            try:
                libro = load_workbook(
                    archivo,
                    data_only=True,
                )

                hoja = libro.active

                encabezados = {}

                for columna, celda in enumerate(
                    hoja[1],
                    start=1,
                ):
                    if celda.value:
                        encabezados[
                            str(celda.value).strip().upper()
                        ] = columna

                columnas_requeridas = [
                    "ITEM",
                    "FECHA RECIBIDO",
                    "HORA RECIBIDO",
                    "ASUNTO",
                    "REMITENTE",
                    "RESPONSABLE",
                    "DEPENDENCIA",
                    "ESTADO INICIAL",
                    "ENVIADO A",
                    "FECHA ASIGNACIÓN",
                    "HORA ASIGNACIÓN",
                    "TERMINO EN DÍAS",
                    "ESTADO ACTUAL",
                    "OBSERVACIONES",
                ]

                faltantes = [
                    columna
                    for columna in columnas_requeridas
                    if columna not in encabezados
                ]

                if faltantes:

                    messages.error(
                        request,
                        "El archivo no contiene las columnas requeridas: "
                        + ", ".join(faltantes)
                    )

                    return render(
                        request,
                        "comunicaciones/importar.html",
                        {
                            "form": form,
                        },
                    )

                creadas = 0
                actualizadas = 0
                errores = []

                for numero_fila in range(
                    2,
                    hoja.max_row + 1,
                ):

                    try:

                        def valor(nombre):
                            columna = encabezados[nombre]

                            return hoja.cell(
                                row=numero_fila,
                                column=columna,
                            ).value

                        item = valor("ITEM")

                        asunto = limpiar_texto(
                            valor("ASUNTO")
                        )

                        remitente = limpiar_texto(
                            valor("REMITENTE")
                        )

                        if not item and not asunto and not remitente:
                            continue

                        try:
                            item_numero = int(item)

                        except (TypeError, ValueError):
                            item_numero = None

                        if item_numero:
                            numero_radicado = (
                                f"COM-2026-{item_numero:04d}"
                            )

                        else:
                            numero_radicado = (
                                f"COM-2026-FILA-{numero_fila}"
                            )

                        termino = valor(
                            "TERMINO EN DÍAS"
                        )

                        try:
                            termino = (
                                int(termino)
                                if termino not in [None, ""]
                                else None
                            )

                        except (TypeError, ValueError):
                            termino = None

                        estado_actual = limpiar_texto(
                            valor("ESTADO ACTUAL")
                        )

                        estado_sistema = (
                            Comunicacion
                            .EstadoComunicacion
                            .PENDIENTE
                        )

                        estado_lower = estado_actual.lower()

                        if "respond" in estado_lower:
                            estado_sistema = (
                                Comunicacion
                                .EstadoComunicacion
                                .RESPONDIDA
                            )

                        elif (
                            "trámite" in estado_lower
                            or "tramite" in estado_lower
                        ):
                            estado_sistema = (
                                Comunicacion
                                .EstadoComunicacion
                                .EN_TRAMITE
                            )

                        elif (
                            "cerr" in estado_lower
                            or "final" in estado_lower
                        ):
                            estado_sistema = (
                                Comunicacion
                                .EstadoComunicacion
                                .CERRADA
                            )

                        comunicacion, creada = (
                            Comunicacion.objects.update_or_create(

                                numero_radicado=numero_radicado,

                                defaults={
                                    "item": item_numero,

                                    "tipo": (
                                        Comunicacion
                                        .TipoComunicacion
                                        .RECIBIDA
                                    ),

                                    "fecha_recibido":
                                        convertir_fecha(
                                            valor(
                                                "FECHA RECIBIDO"
                                            )
                                        ),

                                    "hora_recibido":
                                        convertir_hora(
                                            valor(
                                                "HORA RECIBIDO"
                                            )
                                        ),

                                    "asunto":
                                        asunto,

                                    "remitente":
                                        remitente,

                                    "responsable_texto":
                                        limpiar_texto(
                                            valor(
                                                "RESPONSABLE"
                                            )
                                        ),

                                    "dependencia":
                                        limpiar_texto(
                                            valor(
                                                "DEPENDENCIA"
                                            )
                                        ),

                                    "estado_inicial":
                                        limpiar_texto(
                                            valor(
                                                "ESTADO INICIAL"
                                            )
                                        ),

                                    "enviado_a":
                                        limpiar_texto(
                                            valor(
                                                "ENVIADO A"
                                            )
                                        ),

                                    "fecha_asignacion":
                                        convertir_fecha(
                                            valor(
                                                "FECHA ASIGNACIÓN"
                                            )
                                        ),

                                    "hora_asignacion":
                                        convertir_hora(
                                            valor(
                                                "HORA ASIGNACIÓN"
                                            )
                                        ),

                                    "termino_dias":
                                        termino,

                                    "estado":
                                        estado_sistema,

                                    "estado_actual_texto":
                                        estado_actual,

                                    "observaciones":
                                        limpiar_texto(
                                            valor(
                                                "OBSERVACIONES"
                                            )
                                        ),
                                },
                            )
                        )

                        if creada:
                            creadas += 1
                        else:
                            actualizadas += 1

                    except Exception as error:
                        errores.append(
                            f"Fila {numero_fila}: {error}"
                        )

                resultado = {
                    "creadas": creadas,
                    "actualizadas": actualizadas,
                    "errores": errores,
                }

            except Exception as error:
                messages.error(
                    request,
                    f"No fue posible leer el archivo: {error}",
                )

    else:
        form = ImportarComunicacionesForm()

    return render(
        request,
        "comunicaciones/importar.html",
        {
            "form": form,
            "resultado": resultado,
        },
    )

@login_required
def editar_comunicacion(request, pk):

    comunicacion = get_object_or_404(
        Comunicacion,
        pk=pk,
    )

    if request.method == "POST":

        form = ComunicacionForm(
            request.POST,
            request.FILES,
            instance=comunicacion,
        )

        if form.is_valid():

            comunicacion = form.save()

            # Si fue marcada como respuesta de otra comunicación
            if comunicacion.respuesta_a:

                original = comunicacion.respuesta_a

                original.estado = (
                    Comunicacion
                    .EstadoComunicacion
                    .RESPONDIDA
                )

                original.estado_actual_texto = (
                    "Respondida"
                )

                original.save()

            return redirect(
                "comunicaciones:detalle",
                pk=comunicacion.pk,
            )

    else:

        form = ComunicacionForm(
            instance=comunicacion
        )

    return render(
        request,
        "comunicaciones/editar.html",
        {
            "form": form,
            "comunicacion": comunicacion,
        },
    )